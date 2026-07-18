"""
Routes de génération de rapports financiers et statistiques,
avec export PDF et Excel.
"""
from datetime import datetime, timedelta, timezone
from io import BytesIO

from flask import Blueprint, render_template, request, send_file, current_app
from flask_login import login_required
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill

from app.extensions import db
from app.models.consultation import Consultation
from app.models.payment import Payment
from app.utils.decorators import permission_required
from app.utils.pdf import render_pdf_from_template

reports_bp = Blueprint("reports", __name__, template_folder="../templates/reports")

PERIOD_CHOICES = {
    "jour": "Journalier",
    "semaine": "Hebdomadaire",
    "mois": "Mensuel",
    "annee": "Annuel",
}


def _get_period_bounds(period):
    now = datetime.now(timezone.utc)
    if period == "jour":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "semaine":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "annee":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return start, now


def _build_report_data(period):
    start, end = _get_period_bounds(period)

    payments = Payment.query.filter(Payment.payment_date >= start, Payment.payment_date <= end).all()
    total_revenue = sum(float(p.amount) for p in payments)

    consultations = Consultation.query.filter(
        Consultation.visit_date >= start, Consultation.visit_date <= end, Consultation.is_deleted.is_(False)
    ).all()

    return {
        "period_label": PERIOD_CHOICES.get(period, period),
        "start": start,
        "end": end,
        "total_revenue": total_revenue,
        "payments_count": len(payments),
        "consultations_count": len(consultations),
        "payments": payments,
    }


@reports_bp.route("/")
@login_required
@permission_required("view_reports")
def index():
    period = request.args.get("period", "mois")
    data = _build_report_data(period)
    return render_template("reports/index.html", data=data, period=period, period_choices=PERIOD_CHOICES)


@reports_bp.route("/export/pdf")
@login_required
@permission_required("view_reports")
def export_pdf():
    period = request.args.get("period", "mois")
    data = _build_report_data(period)
    pdf_buffer = render_pdf_from_template(
        "reports/pdf.html", data=data, clinic_name=current_app.config.get("CLINIC_NAME")
    )
    filename = f"rapport_{period}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(pdf_buffer, mimetype="application/pdf", download_name=filename)


@reports_bp.route("/export/excel")
@login_required
@permission_required("view_reports")
def export_excel():
    period = request.args.get("period", "mois")
    data = _build_report_data(period)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rapport"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    sheet.append(["Date du paiement", "Patient", "Montant (DA)", "Méthode", "Reçu par"])
    for col in range(1, 6):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for payment in data["payments"]:
        sheet.append([
            payment.payment_date.strftime("%d/%m/%Y %H:%M"),
            payment.consultation.patient.full_name,
            float(payment.amount),
            payment.method_label,
            payment.received_by.full_name if payment.received_by else "",
        ])

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 4

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"rapport_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=filename,
        as_attachment=True,
    )
